# -*- coding: utf-8 -*-
"""Generate data.js for the Student Report site from the source workbook.
Includes all 5 streams: Bio - Maths, Bio - CS, Maths - CS, Applied Math, and CS.
"""
import openpyxl, json, io, os

SRC = r"C:\Users\balac\Downloads\Students Analysis Report (1).xlsx"
if not os.path.exists(SRC):
    SRC = r"C:\Users\balac\Downloads\Students Analysis Report.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=True)

SUBJECT_FULL = {
    "PHY": "Physics", "CHE": "Chemistry", "MAT": "Mathematics",
    "BIO": "Biology", "CS": "Computer Science", "ENG": "English",
    "PED": "Physical Education", "Acc": "Accountancy", "Bs": "Business Studies",
    "Eco": "Economics", "A.Math": "Applied Mathematics", "Eng": "English",
    "PE": "Physical Education", "Cs": "Computer Science"
}

GROUPS = ["Bio - Maths", "Bio - CS", "Maths - CS", "Applied Math", "CS"]
EXAMS = ["CU 1", "TE 1", "CU 2", "TE 2"]
# Each exam block spans 7 columns (6 subjects + Total). Start column (1-indexed).
BLOCK_START = {"CU 1": 4, "TE 1": 11, "CU 2": 18, "TE 2": 25}

warnings = []

def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def is_absent(v):
    return isinstance(v, str) and v.strip().lower() in ("ab", "absent", "a")

def student_rows(ws):
    rows = []
    for r in range(3, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if b not in (None, "") and is_num(a):
            rows.append(r)
        elif a is not None and not is_num(a):
            break  # reached summary block
    return rows

def midrank_pct(values, x):
    n = len(values)
    if n == 0 or x is None:
        return None
    below = sum(1 for v in values if v < x)
    equal = sum(1 for v in values if v == x)
    return round(100.0 * (below + 0.5 * equal) / n)

# ---------------- PE - Analysis ----------------
wsa = wb["PE - Analysis"]
PE_COLS = {"CU 1": (7, 8), "TE 1": (9, 10), "CU 2": (11, 12), "TE 2": (13, 14)}
pe_students = {}
pe_class = {ex: [] for ex in EXAMS}
for r in student_rows(wsa):
    roll = str(wsa.cell(r, 2).value).strip()
    name = wsa.cell(r, 3).value
    stream = [wsa.cell(r, c).value for c in (4, 5, 6)
              if wsa.cell(r, c).value not in (None, "")]
    ex = {}
    for exname, (tc, rc) in PE_COLS.items():
        tot = wsa.cell(r, tc).value
        rk = wsa.cell(r, rc).value
        tot = tot if is_num(tot) else 0
        rk = rk if is_num(rk) else None
        ex[exname] = {"total": tot, "rank": rk}
        pe_class[exname].append((tot, roll, name))
    pe_students[roll] = {"rollNo": roll, "sNo": wsa.cell(r, 1).value,
                         "name": name, "stream": stream, "exams": ex}

pe_conducted = {}
pe_topper = {}
for ex in EXAMS:
    tots = pe_class[ex]
    mx = max((t[0] for t in tots), default=0)
    pe_conducted[ex] = mx > 0
    if mx > 0:
        top = max(tots, key=lambda t: t[0])
        pe_topper[ex] = {"total": top[0], "roll": top[1], "name": top[2]}
    else:
        pe_topper[ex] = None

# ---------------- Group sheets (subject-wise marks) ----------------
modes = {}
roll_index = {}

def add_index(roll, mode):
    key = roll.upper()
    roll_index.setdefault(key, [])
    if mode not in roll_index[key]:
        roll_index[key].append(mode)

for g in GROUPS:
    ws = wb[g]
    subjects = [str(ws.cell(2, c).value).strip() for c in range(4, 10)]
    rows = student_rows(ws)
    students = {}
    dist = {ex: {s: [] for s in subjects} for ex in EXAMS}
    dist_total = {ex: [] for ex in EXAMS}

    for r in rows:
        roll = str(ws.cell(r, 2).value).strip()
        name = ws.cell(r, 3).value
        marks = {}
        absent = {}
        for ex in EXAMS:
            base = BLOCK_START[ex]
            row_marks = {}
            row_abs = {}
            for i, s in enumerate(subjects):
                v = ws.cell(r, base + i).value
                if is_num(v):
                    row_marks[s] = v
                    dist[ex][s].append(v)
                elif is_absent(v):
                    row_marks[s] = None
                    row_abs[s] = True
                else:
                    row_marks[s] = None
            tot = ws.cell(r, base + 6).value
            row_marks["Total"] = tot if is_num(tot) else 0
            marks[ex] = row_marks
            absent[ex] = row_abs
            if is_num(tot):
                dist_total[ex].append((tot, roll, name))
        students[roll] = {"rollNo": roll, "sNo": ws.cell(r, 1).value,
                          "name": name, "marks": marks, "absent": absent}
        add_index(roll, g)

    conducted = {}
    classStats = {}
    for ex in EXAMS:
        mxtot = max((t[0] for t in dist_total[ex]), default=0)
        conducted[ex] = mxtot > 0
        perSub = {}
        for s in subjects:
            vals = dist[ex][s]
            if vals:
                mx = max(vals)
                toppers = [rl for rl in students
                           if is_num(students[rl]["marks"][ex].get(s))
                           and students[rl]["marks"][ex][s] == mx]
                perSub[s] = {"max": mx, "min": min(vals),
                             "avg": round(sum(vals) / len(vals), 1),
                             "present": len(vals),
                             "topperRoll": toppers[0] if toppers else None,
                             "topperName": students[toppers[0]]["name"] if toppers else None}
            else:
                perSub[s] = None
        if dist_total[ex]:
            top = max(dist_total[ex], key=lambda t: t[0])
            totStat = {"max": top[0], "topperRoll": top[1], "topperName": top[2],
                       "present": len(dist_total[ex])}
        else:
            totStat = None
        classStats[ex] = {"subjects": perSub, "total": totStat}

    for roll, st in students.items():
        st["percentile"] = {}
        for ex in EXAMS:
            if not conducted[ex]:
                continue
            pex = {}
            for s in subjects:
                mk = st["marks"][ex].get(s)
                pex[s] = midrank_pct(dist[ex][s], mk) if is_num(mk) else None
            tot = st["marks"][ex].get("Total")
            pex["Total"] = midrank_pct([t[0] for t in dist_total[ex]], tot) if is_num(tot) else None
            st["percentile"][ex] = pex
        pe = pe_students.get(roll)
        st["overall"] = {ex: (pe["exams"][ex] if pe else None) for ex in EXAMS} if pe else None

    modes[g] = {"type": "group", "label": g, "subjects": subjects,
                "subjectFull": {s: SUBJECT_FULL.get(s, s) for s in subjects},
                "exams": EXAMS, "conducted": conducted,
                "classSize": len(students), "classStats": classStats,
                "students": students}

for roll in pe_students:
    add_index(roll, "PE - Analysis")

modes["PE - Analysis"] = {"type": "analysis", "label": "PE - Analysis",
                          "exams": EXAMS, "conducted": pe_conducted,
                          "classSize": len(pe_students), "topper": pe_topper,
                          "students": pe_students}

data = {
    "meta": {"source": "Students Analysis Report (1).xlsx",
             "academicYear": "2026 - 2027",
             "maxPerSubject": 100,
             "note": "Only CU 1 has been conducted; TE 1 / CU 2 / TE 2 are pending."},
    "modeOrder": ["PE - Analysis", "Bio - Maths", "Bio - CS", "Maths - CS", "Applied Math", "CS"],
    "modes": modes,
    "rollIndex": roll_index,
}

with io.open("data.js", "w", encoding="utf-8") as f:
    f.write("/* AUTO-GENERATED from 'Students Analysis Report (1).xlsx'. Do not edit by hand. */\n")
    f.write("window.REPORT_DATA = " + json.dumps(data, ensure_ascii=False, indent=1) + ";\n")

print("WROTE data.js")
print("Class sizes:", {m: modes[m]["classSize"] for m in modes})
print("Unique rolls indexed:", len(roll_index))
